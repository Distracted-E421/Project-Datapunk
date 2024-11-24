from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
from datetime import datetime, timedelta
from enum import Enum
import json
import asyncio
from .monitoring import HealthMetrics, MonitoringLevel
from ..discovery.registry import ServiceRegistration
import pandas as pd
import numpy as np
from pathlib import Path
import csv
import io
from jinja2 import Template
import tabulate
import base64
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font

class ReportFormat(Enum):
    """Supported report formats"""
    JSON = "json"
    HTML = "html"
    CSV = "csv"
    MARKDOWN = "markdown"
    EXCEL = "excel"

class ReportType(Enum):
    """Types of health reports"""
    SUMMARY = "summary"
    DETAILED = "detailed"
    METRICS = "metrics"
    ALERTS = "alerts"
    TRENDS = "trends"

@dataclass
class ReportConfig:
    """Configuration for health reporting"""
    report_dir: str = "./reports"
    default_format: ReportFormat = ReportFormat.JSON
    retention_days: int = 30
    include_metrics: bool = True
    include_alerts: bool = True
    include_trends: bool = True
    max_trend_points: int = 100
    auto_cleanup: bool = True

class HealthReporter:
    """Handles health report generation and formatting"""
    def __init__(
        self,
        config: ReportConfig,
        metrics_collector: Optional[MetricsCollector] = None
    ):
        self.config = config
        self.metrics = metrics_collector
        Path(config.report_dir).mkdir(parents=True, exist_ok=True)

    async def generate_report(
        self,
        report_type: ReportType,
        format: Optional[ReportFormat] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        services: Optional[List[str]] = None
    ) -> Union[str, Dict]:
        """Generate health report"""
        format = format or self.config.default_format
        end_time = end_time or datetime.utcnow()
        start_time = start_time or (end_time - timedelta(hours=24))

        try:
            # Generate base report data
            if report_type == ReportType.SUMMARY:
                report_data = await self._generate_summary_report(services)
            elif report_type == ReportType.DETAILED:
                report_data = await self._generate_detailed_report(services)
            elif report_type == ReportType.METRICS:
                report_data = await self._generate_metrics_report(start_time, end_time, services)
            elif report_type == ReportType.ALERTS:
                report_data = await self._generate_alerts_report(start_time, end_time, services)
            elif report_type == ReportType.TRENDS:
                report_data = await self._generate_trends_report(services)
            else:
                raise ValueError(f"Unsupported report type: {report_type}")

            # Format report
            formatted_report = await self._format_report(report_data, format)

            # Save report if directory configured
            if self.config.report_dir:
                await self._save_report(formatted_report, report_type, format)

            if self.metrics:
                await self.metrics.increment(
                    "health.report.generated",
                    tags={
                        "type": report_type.value,
                        "format": format.value
                    }
                )

            return formatted_report

        except Exception as e:
            if self.metrics:
                await self.metrics.increment(
                    "health.report.error",
                    tags={
                        "type": report_type.value,
                        "error": str(e)
                    }
                )
            raise

    async def _generate_summary_report(
        self,
        services: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Generate summary health report"""
        report = {
            "timestamp": datetime.utcnow().isoformat(),
            "services": {},
            "overall_status": "healthy",
            "metrics": {
                "total_services": 0,
                "healthy_services": 0,
                "degraded_services": 0,
                "unhealthy_services": 0
            }
        }

        # Collect service health data
        for service_id in (services or self._services.keys()):
            service_metrics = await self._get_service_metrics(service_id)
            if not service_metrics:
                continue

            latest_metric = service_metrics[-1]
            report["services"][service_id] = {
                "status": latest_metric.status.value,
                "error_rate": latest_metric.error_rate,
                "response_time": latest_metric.response_time,
                "last_check": latest_metric.last_check.isoformat()
            }

            # Update metrics
            report["metrics"]["total_services"] += 1
            if latest_metric.status == HealthStatus.HEALTHY:
                report["metrics"]["healthy_services"] += 1
            elif latest_metric.status == HealthStatus.DEGRADED:
                report["metrics"]["degraded_services"] += 1
            else:
                report["metrics"]["unhealthy_services"] += 1

        # Determine overall status
        if report["metrics"]["unhealthy_services"] > 0:
            report["overall_status"] = "unhealthy"
        elif report["metrics"]["degraded_services"] > 0:
            report["overall_status"] = "degraded"

        return report

    async def _generate_detailed_report(
        self,
        services: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Generate detailed health report"""
        report = {
            "timestamp": datetime.utcnow().isoformat(),
            "services": {},
            "alerts": [],
            "trends": {}
        }

        for service_id in (services or self._services.keys()):
            service_metrics = await self._get_service_metrics(service_id)
            if not service_metrics:
                continue

            latest_metric = service_metrics[-1]
            report["services"][service_id] = {
                "current_status": latest_metric.status.value,
                "metrics": {
                    "error_rate": {
                        "current": latest_metric.error_rate,
                        "threshold": self.config.error_threshold
                    },
                    "response_time": {
                        "current": latest_metric.response_time,
                        "threshold": self.config.response_time_threshold
                    },
                    "resource_usage": {
                        "memory": latest_metric.memory_usage,
                        "cpu": latest_metric.cpu_usage,
                        "connections": latest_metric.connections
                    }
                },
                "last_check": latest_metric.last_check.isoformat()
            }

            # Add alerts
            alerts = await self._get_service_alerts(service_id)
            report["alerts"].extend(alerts)

            # Add trends
            report["trends"][service_id] = await self._calculate_trends(service_metrics)

        return report

    async def _save_report(
        self,
        report: Union[str, Dict],
        report_type: ReportType,
        format: ReportFormat
    ):
        """Save report to file"""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type.value}_{timestamp}.{format.value}"
        filepath = Path(self.config.report_dir) / filename

        try:
            if isinstance(report, dict):
                with open(filepath, 'w') as f:
                    json.dump(report, f, indent=2)
            else:
                with open(filepath, 'w') as f:
                    f.write(report)

            # Cleanup old reports if enabled
            if self.config.auto_cleanup:
                await self._cleanup_old_reports()

        except Exception as e:
            if self.metrics:
                await self.metrics.increment(
                    "health.report.save_error",
                    tags={"error": str(e)}
                )
            raise

    async def _cleanup_old_reports(self):
        """Clean up old reports"""
        if not self.config.retention_days:
            return

        cutoff_time = datetime.utcnow() - timedelta(days=self.config.retention_days)
        report_dir = Path(self.config.report_dir)

        for report_file in report_dir.glob("*.*"):
            try:
                # Extract timestamp from filename
                timestamp_str = report_file.stem.split("_")[-2:]
                timestamp = datetime.strptime("_".join(timestamp_str), "%Y%m%d_%H%M%S")

                if timestamp < cutoff_time:
                    report_file.unlink()

            except (ValueError, IndexError):
                continue  # Skip files that don't match expected format

    async def _calculate_trends(
        self,
        metrics: List[HealthMetrics]
    ) -> Dict[str, Any]:
        """Calculate trends from metrics"""
        if not metrics:
            return {}

        # Calculate moving averages
        window_size = min(len(metrics), 10)
        error_rates = [m.error_rate for m in metrics]
        response_times = [m.response_time for m in metrics]

        return {
            "error_rate": {
                "trend": self._calculate_moving_average(error_rates, window_size),
                "direction": self._calculate_trend_direction(error_rates)
            },
            "response_time": {
                "trend": self._calculate_moving_average(response_times, window_size),
                "direction": self._calculate_trend_direction(response_times)
            }
        }

    def _calculate_moving_average(
        self,
        values: List[float],
        window_size: int
    ) -> List[float]:
        """Calculate moving average of values"""
        result = []
        for i in range(len(values)):
            window_start = max(0, i - window_size + 1)
            window = values[window_start:i + 1]
            result.append(sum(window) / len(window))
        return result

    def _calculate_trend_direction(self, values: List[float]) -> str:
        """Calculate trend direction (improving, stable, degrading)"""
        if len(values) < 2:
            return "stable"

        recent_avg = sum(values[-3:]) / min(3, len(values))
        older_avg = sum(values[:-3]) / max(1, len(values) - 3)

        diff = recent_avg - older_avg
        if abs(diff) < 0.05:  # 5% threshold
            return "stable"
        return "improving" if diff < 0 else "degrading"

    async def _generate_metrics_report(
        self,
        metrics: Dict[str, List[HealthMetrics]]
    ) -> Dict[str, Any]:
        """Generate metrics-focused report"""
        report = {
            "timestamp": datetime.utcnow().isoformat(),
            "metrics": {}
        }

        for service_id, service_metrics in metrics.items():
            if not service_metrics:
                continue

            df = pd.DataFrame([vars(m) for m in service_metrics])
            
            report["metrics"][service_id] = {
                "error_rate": {
                    "current": float(df["error_rate"].iloc[-1]),
                    "mean": float(df["error_rate"].mean()),
                    "std": float(df["error_rate"].std()),
                    "percentiles": {
                        "50": float(df["error_rate"].quantile(0.5)),
                        "90": float(df["error_rate"].quantile(0.9)),
                        "95": float(df["error_rate"].quantile(0.95)),
                        "99": float(df["error_rate"].quantile(0.99))
                    }
                },
                "response_time": {
                    "current": float(df["response_time"].iloc[-1]),
                    "mean": float(df["response_time"].mean()),
                    "std": float(df["response_time"].std()),
                    "percentiles": {
                        "50": float(df["response_time"].quantile(0.5)),
                        "90": float(df["response_time"].quantile(0.9)),
                        "95": float(df["response_time"].quantile(0.95)),
                        "99": float(df["response_time"].quantile(0.99))
                    }
                }
            }

        return report

    async def _generate_alerts_report(
        self,
        metrics: Dict[str, List[HealthMetrics]]
    ) -> Dict[str, Any]:
        """Generate alerts report"""
        report = {
            "timestamp": datetime.utcnow().isoformat(),
            "alerts": []
        }

        for service_id, service_metrics in metrics.items():
            if not service_metrics:
                continue

            latest = service_metrics[-1]
            
            # Check various alert conditions
            if latest.error_rate > 0.1:
                report["alerts"].append({
                    "service": service_id,
                    "type": "error_rate",
                    "level": MonitoringLevel.WARNING.value,
                    "message": f"High error rate: {latest.error_rate*100}%",
                    "timestamp": latest.last_check.isoformat()
                })

            if latest.memory_usage > 0.9:
                report["alerts"].append({
                    "service": service_id,
                    "type": "memory",
                    "level": MonitoringLevel.CRITICAL.value,
                    "message": f"High memory usage: {latest.memory_usage*100}%",
                    "timestamp": latest.last_check.isoformat()
                })

            if latest.cpu_usage > 0.8:
                report["alerts"].append({
                    "service": service_id,
                    "type": "cpu",
                    "level": MonitoringLevel.WARNING.value,
                    "message": f"High CPU usage: {latest.cpu_usage*100}%",
                    "timestamp": latest.last_check.isoformat()
                })

        return report

    async def _generate_trends_report(
        self,
        metrics: Dict[str, List[HealthMetrics]]
    ) -> Dict[str, Any]:
        """Generate trends report"""
        report = {
            "timestamp": datetime.utcnow().isoformat(),
            "trends": {}
        }

        for service_id, service_metrics in metrics.items():
            if not service_metrics:
                continue

            df = pd.DataFrame([vars(m) for m in service_metrics])
            
            # Calculate rolling averages
            window = min(len(df), 10)  # 10-point rolling average
            report["trends"][service_id] = {
                "error_rate": {
                    "current": df["error_rate"].tolist()[-self.config.max_trend_points:],
                    "trend": df["error_rate"].rolling(window=window).mean().tolist()[-self.config.max_trend_points:]
                },
                "response_time": {
                    "current": df["response_time"].tolist()[-self.config.max_trend_points:],
                    "trend": df["response_time"].rolling(window=window).mean().tolist()[-self.config.max_trend_points:]
                },
                "resource_usage": {
                    "memory": df["memory_usage"].tolist()[-self.config.max_trend_points:],
                    "cpu": df["cpu_usage"].tolist()[-self.config.max_trend_points:]
                }
            }

        return report

    async def _format_report(
        self,
        report_data: Dict[str, Any],
        format: ReportFormat
    ) -> Union[str, Dict]:
        """Format report in specified format"""
        if format == ReportFormat.JSON:
            return report_data
        elif format == ReportFormat.CSV:
            return self._to_csv(report_data)
        elif format == ReportFormat.HTML:
            return self._to_html(report_data)
        elif format == ReportFormat.MARKDOWN:
            return self._to_markdown(report_data)
        elif format == ReportFormat.EXCEL:
            return self._to_excel(report_data)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _filter_metrics_by_time(
        self,
        metrics: Dict[str, List[HealthMetrics]],
        start_time: datetime,
        end_time: datetime
    ) -> Dict[str, List[HealthMetrics]]:
        """Filter metrics by time range"""
        filtered = {}
        for service_id, service_metrics in metrics.items():
            filtered[service_id] = [
                m for m in service_metrics
                if start_time <= m.last_check <= end_time
            ]
        return filtered

    # Format conversion methods (implement based on your needs)
    def _to_csv(self, data: Dict) -> str:
        """Convert report to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(["Timestamp", data["timestamp"]])
        writer.writerow([])  # Empty row for spacing
        
        if "services" in data:
            writer.writerow(["Services"])
            writer.writerow(["ID", "Status", "Error Rate", "Response Time", "Memory", "CPU", "Connections"])
            
            for service_id, info in data["services"].items():
                writer.writerow([
                    service_id,
                    info.get("current_status", "unknown"),
                    f"{info.get('metrics', {}).get('error_rate', {}).get('current', 0)*100:.2f}%",
                    f"{info.get('metrics', {}).get('response_time', {}).get('current', 0):.3f}s",
                    f"{info.get('metrics', {}).get('resource_usage', {}).get('memory', 0)*100:.1f}%",
                    f"{info.get('metrics', {}).get('resource_usage', {}).get('cpu', 0)*100:.1f}%",
                    info.get('metrics', {}).get('resource_usage', {}).get('connections', 0)
                ])
        
        if "alerts" in data:
            writer.writerow([])
            writer.writerow(["Alerts"])
            writer.writerow(["Service", "Type", "Level", "Message", "Timestamp"])
            
            for alert in data["alerts"]:
                writer.writerow([
                    alert["service"],
                    alert["type"],
                    alert["level"],
                    alert["message"],
                    alert["timestamp"]
                ])
        
        return output.getvalue()

    def _to_html(self, data: Dict) -> str:
        """Convert report to HTML format"""
        template = Template("""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Health Report - {{ timestamp }}</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                th, td { padding: 8px; text-align: left; border: 1px solid #ddd; }
                th { background-color: #f5f5f5; }
                .healthy { color: green; }
                .unhealthy { color: red; }
                .warning { color: orange; }
                .critical { color: darkred; }
                .metrics { margin: 20px 0; }
                .alerts { margin: 20px 0; background-color: #fff3f3; padding: 10px; }
            </style>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
            <script>
            function createChart(canvasId, labels, data, label) {
                const ctx = document.getElementById(canvasId).getContext('2d');
                new Chart(ctx, {
                    type: 'line',
                    data: {
                        labels: labels,
                        datasets: [{
                            label: label,
                            data: data,
                            borderColor: 'rgb(75, 192, 192)',
                            tension: 0.1
                        }]
                    },
                    options: {
                        responsive: true,
                        scales: {
                            y: {
                                beginAtZero: true
                            }
                        }
                    }
                });
            }
            </script>
        </head>
        <body>
            <h1>Health Report</h1>
            <p>Generated at: {{ timestamp }}</p>
            
            {% if services %}
            <div class="metrics">
                <h2>Services</h2>
                <table>
                    <tr>
                        <th>Service ID</th>
                        <th>Status</th>
                        <th>Error Rate</th>
                        <th>Response Time</th>
                        <th>Memory Usage</th>
                        <th>CPU Usage</th>
                        <th>Connections</th>
                    </tr>
                    {% for service_id, info in services.items() %}
                    <tr>
                        <td>{{ service_id }}</td>
                        <td class="{{ info.current_status }}">{{ info.current_status }}</td>
                        <td>{{ "%.2f"|format(info.metrics.error_rate.current * 100) }}%</td>
                        <td>{{ "%.3f"|format(info.metrics.response_time.current) }}s</td>
                        <td>{{ "%.1f"|format(info.metrics.resource_usage.memory * 100) }}%</td>
                        <td>{{ "%.1f"|format(info.metrics.resource_usage.cpu * 100) }}%</td>
                        <td>{{ info.metrics.resource_usage.connections }}</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% endif %}
            
            {% if alerts %}
            <div class="alerts">
                <h2>Alerts</h2>
                <table>
                    <tr>
                        <th>Service</th>
                        <th>Type</th>
                        <th>Level</th>
                        <th>Message</th>
                        <th>Timestamp</th>
                    </tr>
                    {% for alert in alerts %}
                    <tr>
                        <td>{{ alert.service }}</td>
                        <td>{{ alert.type }}</td>
                        <td class="{{ alert.level }}">{{ alert.level }}</td>
                        <td>{{ alert.message }}</td>
                        <td>{{ alert.timestamp }}</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% endif %}
            
            {% if trends %}
            <div class="trends">
                <h2>Trends</h2>
                {% for service_id, trend in trends.items() %}
                <h3>{{ service_id }}</h3>
                <div class="trend-chart">
                    <canvas id="trend-{{ service_id }}-error_rate"></canvas>
                    <script>
                        createChart(
                            'trend-{{ service_id }}-error_rate',
                            {{ range(trend_data|length)|list|tojson }},
                            {{ trend_data.error_rate.current|tojson }},
                            'Error Rate'
                        );
                    </script>
                </div>
                <div class="trend-chart">
                    <canvas id="trend-{{ service_id }}-response_time"></canvas>
                    <script>
                        createChart(
                            'trend-{{ service_id }}-response_time',
                            {{ range(trend_data|length)|list|tojson }},
                            {{ trend_data.response_time.current|tojson }},
                            'Response Time'
                        );
                    </script>
                </div>
                <div class="trend-chart">
                    <canvas id="trend-{{ service_id }}-memory"></canvas>
                    <script>
                        createChart(
                            'trend-{{ service_id }}-memory',
                            {{ range(trend_data|length)|list|tojson }},
                            {{ trend_data.resource_usage.memory|tojson }},
                            'Memory'
                        );
                    </script>
                </div>
                <div class="trend-chart">
                    <canvas id="trend-{{ service_id }}-cpu"></canvas>
                    <script>
                        createChart(
                            'trend-{{ service_id }}-cpu',
                            {{ range(trend_data|length)|list|tojson }},
                            {{ trend_data.resource_usage.cpu|tojson }},
                            'CPU'
                        );
                    </script>
                </div>
                {% endfor %}
            </div>
            {% endif %}
        </body>
        </html>
        """)
        
        return template.render(**data)

    def _to_markdown(self, data: Dict) -> str:
        """Convert report to Markdown format"""
        output = []
        
        # Header
        output.append(f"# Health Report\n")
        output.append(f"Generated at: {data['timestamp']}\n")
        
        # Services
        if "services" in data:
            output.append("## Services\n")
            
            # Prepare table data
            headers = ["Service ID", "Status", "Error Rate", "Response Time", "Memory", "CPU", "Connections"]
            table_data = []
            
            for service_id, info in data["services"].items():
                table_data.append([
                    service_id,
                    info.get("current_status", "unknown"),
                    f"{info.get('metrics', {}).get('error_rate', {}).get('current', 0)*100:.2f}%",
                    f"{info.get('metrics', {}).get('response_time', {}).get('current', 0):.3f}s",
                    f"{info.get('metrics', {}).get('resource_usage', {}).get('memory', 0)*100:.1f}%",
                    f"{info.get('metrics', {}).get('resource_usage', {}).get('cpu', 0)*100:.1f}%",
                    info.get('metrics', {}).get('resource_usage', {}).get('connections', 0)
                ])
            
            output.append(tabulate.tabulate(table_data, headers=headers, tablefmt="pipe"))
            output.append("\n")
        
        # Alerts
        if "alerts" in data:
            output.append("## Alerts\n")
            
            headers = ["Service", "Type", "Level", "Message", "Timestamp"]
            table_data = [
                [
                    alert["service"],
                    alert["type"],
                    alert["level"],
                    alert["message"],
                    alert["timestamp"]
                ]
                for alert in data["alerts"]
            ]
            
            output.append(tabulate.tabulate(table_data, headers=headers, tablefmt="pipe"))
            output.append("\n")
        
        # Trends
        if "trends" in data:
            output.append("## Trends\n")
            
            for service_id, trend_data in data["trends"].items():
                output.append(f"### {service_id}\n")
                output.append("```\n")
                # Add ASCII chart representation here if needed
                output.append("```\n")
        
        return "\n".join(output)

    def _to_excel(self, data: Dict) -> bytes:
        """Convert report to Excel format"""
        wb = Workbook()
        
        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Health Report", data["timestamp"]])
        ws.append([])
        
        # Services sheet
        if "services" in data:
            services_sheet = wb.create_sheet("Services")
            headers = ["Service ID", "Status", "Error Rate", "Response Time", "Memory", "CPU", "Connections"]
            services_sheet.append(headers)
            
            # Style headers
            for cell in services_sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="CCCCCC")
            
            for service_id, info in data["services"].items():
                row = [
                    service_id,
                    info.get("current_status", "unknown"),
                    info.get('metrics', {}).get('error_rate', {}).get('current', 0),
                    info.get('metrics', {}).get('response_time', {}).get('current', 0),
                    info.get('metrics', {}).get('resource_usage', {}).get('memory', 0),
                    info.get('metrics', {}).get('resource_usage', {}).get('cpu', 0),
                    info.get('metrics', {}).get('resource_usage', {}).get('connections', 0)
                ]
                services_sheet.append(row)
                
                # Add conditional formatting
                status_cell = services_sheet.cell(services_sheet.max_row, 2)
                if row[1] == "healthy":
                    status_cell.fill = PatternFill("solid", fgColor="90EE90")
                elif row[1] == "unhealthy":
                    status_cell.fill = PatternFill("solid", fgColor="FFB6C1")
            
            # Add charts
            if "trends" in data:
                trends_sheet = wb.create_sheet("Trends")
                for service_id, trend_data in data["trends"].items():
                    # Add error rate trend
                    chart = LineChart()
                    chart.title = f"{service_id} - Error Rate Trend"
                    chart.y_axis.title = "Error Rate"
                    chart.x_axis.title = "Time"
                    
                    trends_sheet.append([f"{service_id} Error Rate"])
                    data_rows = trend_data["error_rate"]["current"]
                    for i, value in enumerate(data_rows):
                        trends_sheet.append([i, value])
                    
                    data = Reference(trends_sheet, min_col=2, min_row=1, max_row=len(data_rows))
                    chart.add_data(data, titles_from_data=True)
                    trends_sheet.add_chart(chart, "A10")
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _format_trend_ascii(self, values: List[float], width: int = 50, height: int = 10) -> str:
        """Create enhanced ASCII chart for trend visualization"""
        if not values:
            return ""
        
        # Normalize values to height
        min_val = min(values)
        max_val = max(values)
        range_val = max_val - min_val or 1
        
        normalized = [
            int((v - min_val) / range_val * (height - 1))
            for v in values
        ]
        
        # Create chart with axis labels
        chart = []
        
        # Add max value label
        chart.append(f"{max_val:8.2f} ┤")
        
        # Create chart body
        for y in range(height - 1, -1, -1):
            # Add y-axis label every few lines
            if y % (height // 4) == 0:
                current_val = min_val + (y / (height - 1)) * range_val
                prefix = f"{current_val:8.2f} ┤"
            else:
                prefix = "         │"
            
            line = [
                "█" if n >= y else "·" if n == y - 1 else " "
                for n in normalized
            ]
            chart.append(prefix + "".join(line))
        
        # Add x-axis
        chart.append("         └" + "─" * len(values))
        
        # Add x-axis labels
        x_labels = []
        step = max(len(values) // 5, 1)
        for i in range(0, len(values), step):
            x_labels.append(str(i).center(step))
        chart.append("          " + "".join(x_labels))
        
        return "\n".join(chart)

    def _generate_svg_chart(self, values: List[float], title: str) -> str:
        """Generate SVG chart for Markdown"""
        width = 600
        height = 300
        padding = 40
        
        # Normalize values
        min_val = min(values)
        max_val = max(values)
        range_val = max_val - min_val or 1
        
        # Generate points
        points = []
        x_step = (width - 2 * padding) / (len(values) - 1)
        for i, v in enumerate(values):
            x = padding + i * x_step
            y = height - padding - ((v - min_val) / range_val) * (height - 2 * padding)
            points.append(f"{x},{y}")
        
        return f"""
        <svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg">
            <style>
                .title {{ font-size: 16px; font-family: sans-serif; }}
                .axis {{ stroke: #ccc; stroke-width: 1; }}
                .line {{ stroke: #2196F3; stroke-width: 2; fill: none; }}
            </style>
            <text x="{width/2}" y="20" text-anchor="middle" class="title">{title}</text>
            <line x1="{padding}" y1="{height-padding}" x2="{width-padding}" y2="{height-padding}" class="axis" />
            <line x1="{padding}" y1="{padding}" x2="{padding}" y2="{height-padding}" class="axis" />
            <polyline points="{' '.join(points)}" class="line" />
        </svg>
        """ 